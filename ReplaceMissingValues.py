from flask import Flask, request, render_template, send_file
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

def process_file(file_path):
    df = pd.read_excel(file_path)

    df.columns = ['Person ID', 'Name', 'Date', 'Attendance Status', 'Check-In', 'Check-Out']

    df['Check-In'] = pd.to_datetime(df['Check-In'], format='%H:%M:%S', errors='coerce').dt.time
    df['Check-Out'] = pd.to_datetime(df['Check-Out'], format='%H:%M:%S', errors='coerce').dt.time
    df['Attendance Status'].fillna('Normal', inplace=True)

    def average_time(times):
        total_seconds = ([t.hour * 3600 + t.minute * 60 + t.second for t in times if t is not pd.NaT])
        if len(total_seconds) == 0:
            return pd.NaT
        avg_seconds = round(sum(total_seconds) / len(total_seconds))
        avg_time = (datetime.min + timedelta(seconds=avg_seconds)).time()
        return avg_time

    average_checkin_per_person = df.groupby('Name')['Check-In'].apply(lambda x: average_time(x.tolist()))
    average_checkout_per_person = df.groupby('Name')['Check-Out'].apply(lambda x: average_time(x.tolist()))

    checkin_fill_color = []
    checkout_fill_color = []

    df['Check-In'] = df.apply(
        lambda row: (
            average_checkin_per_person[row['Name']], checkin_fill_color.append(True)
        ) if pd.isna(row['Check-In']) else (row['Check-In'], checkin_fill_color.append(False)),
        axis=1
    )

    df['Check-Out'] = df.apply(
        lambda row: (
            average_checkout_per_person[row['Name']], checkout_fill_color.append(True)
        ) if pd.isna(row['Check-Out']) else (row['Check-Out'], checkout_fill_color.append(False)),
        axis=1
    )

    df[['Check-In', 'Check-Out']] = df[['Check-In', 'Check-Out']].applymap(lambda x: x[0] if isinstance(x, tuple) else x)

    output_path = 'absen_full.xlsx'
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

    for idx, fill in enumerate(checkin_fill_color):
        if fill:
            ws.cell(row=idx + 2, column=5).fill = red_fill

    for idx, fill in enumerate(checkout_fill_color):
        if fill:
            ws.cell(row=idx + 2, column=6).fill = red_fill

    wb.save(output_path)

    return output_path

@app.route('/')
def upload_file():
    return render_template('upload.html')

@app.route('/uploader', methods=['GET', 'POST'])
def upload_file_post():
    if request.method == 'POST':
        f = request.files['file']
        file_path = os.path.join('uploads', f.filename)
        f.save(file_path)
        output_path = process_file(file_path)
        return send_file(output_path, as_attachment=True)

if __name__ == '__main__':
    if not os.path.exists('uploads'):
        os.makedirs('uploads')
    app.run(debug=True, host="0.0.0.0")
